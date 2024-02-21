#------------------------------------------------------------------------------------------------------------------------------
#------------------------------------------------------------------------------------------------------------------------------
#---------Código Realizado por Pablo Castro, practicante entre Enero-Marzo 2024 para banco Santander.--------------------------
#------------------------------------------------------------------------------------------------------------------------------
#--------------------------------------------------------Contacto: pablocastro21@gmail.com ------------------------------------
#------------------------------------------------------------------------------------------------------------------------------
#------------------------------------------------------------------------------------------------------------------------------

import openpyxl
import os
from openpyxl import worksheet
from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from tkinter.simpledialog import askinteger
from openpyxl.styles import Alignment, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy
import pandas as pd
import numpy as np
from itertools import combinations

relleno_destacado = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def importar_y_obtener_dataframe(archivo_excel, nombre_hoja):
    wb = openpyxl.load_workbook(archivo_excel, data_only=True)
    hoja = wb[nombre_hoja]
    datos = hoja.iter_rows(min_row=7, max_col=12, values_only=True)
    dataframe = pd.DataFrame(datos, columns=[openpyxl.utils.get_column_letter(col) for col in range(1, 13)])
    return dataframe

def calcular_sumas_por_columna(df):
    sumas = []
    for columna in df.columns[1:]:  # Excluir la primera columna
        suma_columna = df[columna].sum()
        sumas.append(suma_columna)
    return sumas

def optimizar(ruta, hoja):
    df = importar_y_obtener_dataframe(ruta, hoja) #aqui esta el df desde fila 7, columna A hasta L
    indices_filas = list(df.index) #índices de las filas del df
    combinaciones_validas = []
    df_filtrado = pd.DataFrame()
    a=0
    b=0
    tamaño_combinacion = 2
    print(df.columns)
     
    for combinacion in combinations(indices_filas, tamaño_combinacion): #combinacion de indices de filas
        df_combinacion = df.loc[list(combinacion)]  # Utilizar .loc para subconjunto por índices
        #Con estos genera un df nuevo, que tiene las filas de la combinacion
        sumas_combinacion = calcular_sumas_por_columna(df_combinacion)
        #calcula las sumas de las columnas, de este nuevo df, excluyendo A.
        #Retorna lista con las sumas, sin la columna A.
        print ('combinacion:', a)
        print ('Pares encontrados: ', b)
           
        if all(abs(sc) < 20 for sc in (sumas_combinacion)) and not any(idx in df_filtrado.index for idx in combinacion):
            ids_filas = df_combinacion['A'].values # Asumiendo 'A' como columna de IDs
            combinaciones_validas.append((ids_filas, sumas_combinacion))
            df_filtrado = pd.concat([df_filtrado, df_combinacion])
            print(f"Combinación válida encontrada: IDs de filas {ids_filas},Sumas {sumas_combinacion}")
            b+=1
            if b==250 : 
                print(f"Total de combinaciones válidas encontradas: {len(combinaciones_validas)}")
                break  
        a+=1 
            
    # Después de revisar todas las combinaciones:
    print(f"Total de combinaciones válidas encontradas: {len(combinaciones_validas)}")
    return(df_filtrado)

def exportar_tabla(df, archivo_original):
    nuevo_nombre = archivo_original.replace('.xlsx', '') + '_MOD.xlsx'
    df.to_excel(nuevo_nombre, index=False)
    # Abrir el archivo nuevo y aplicar el formato
    libro = openpyxl.load_workbook(nuevo_nombre)
    hoja_original = libro.active

    # Crear las hojas adicionales
    hoja_filas_eliminadas = libro.create_sheet("Filas Eliminadas")
    hoja_datos_finales = libro.create_sheet("Datos Finales")

    # Copiar las primeras 6 filas de la hoja original a las nuevas hojas y aplicar formato
    for i in range(1, 7):
        for col in range(1, len(df.columns) + 1):
            for hoja_destino in [hoja_filas_eliminadas, hoja_datos_finales]:
                celda_original = hoja_original.cell(row=i, column=col)
                celda_nueva = hoja_destino.cell(row=i, column=col, value=celda_original.value)

                # Aplicar el mismo relleno a las filas 4 y 6
                if i in [4, 6]:
                    celda_original.fill = relleno_destacado
                    celda_nueva.fill = relleno_destacado
            # Aquí puedes copiar otros atributos de la celda, como formato, si es necesario

    # Aplicar la función quitarnegativos a la hoja clon

    # Guardar los cambios en el libro de trabajo
    libro.save(nuevo_nombre)
    libro.close()
    print(f"Archivo exportado como {nuevo_nombre}")
    ruta_archivo = os.path.abspath(nuevo_nombre)
    return ruta_archivo

def analizar_excel(archivo, columnas):
    # Lee y extrae la primera hoja, la tabla entre columnas indicadas.
    df = pd.read_excel(archivo, sheet_name=0, usecols=columnas)
    # Rellenar los valores NaN con 0
    df.fillna(0, inplace=True)
    return df

def seleccionar_archivo_excel():
    Tk().withdraw()  # no queremos una GUI completa, solo la caja de diálogo
    archivo = askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])
    return archivo

def inyectar_df_en_excel(df, ruta_excel):
    # Cargar el libro de Excel
    libro = load_workbook(ruta_excel)

    # Seleccionar la hoja 'Datos Finales'
    hoja_datos_finales = libro['Datos Finales']
    hoja_filas_eliminadas = libro['Filas Eliminadas']
    hoja_sheet_1 = libro['Sheet1']
    ids_df = set(df['A'])

    # Determinar la fila y columna de inicio
    fila_inicio = 7
    columna_inicio = 1  # La columna A equivale a 1

    # Iterar sobre el dataframe y escribir cada valor en la hoja, una fila después de la otra
    for indice, fila in df.iterrows():
        for j, valor in enumerate(fila):
            celda = hoja_datos_finales.cell(row=fila_inicio + indice, column=columna_inicio + j)
            celda.value = valor
    
    for f in hoja_sheet_1.iter_rows(min_row=7, values_only = True):
        id_fila = f[0]

        if id_fila not in ids_df:
            fila_a_copiar = list(f)
            for r in dataframe_to_rows(pd.DataFrame([fila_a_copiar]), index=False, header=False):
                hoja_filas_eliminadas.append(r)

            fila_inicio+=1
    for sheet in libro.worksheets:
        # Set the range for the sum formula
        start_row = 7
        end_row = sheet.max_row
        start_col = 2
        end_col = 12
        # Set the formula in each cell of row 4 from column B to column L
        for col in range(start_col, end_col + 1):
            sum_formula = '=SUM({}{}:{}{})'.format(chr(col + 64), start_row, chr(col + 64), end_row)
            sheet.cell(row=4, column=col, value=sum_formula)
    # Remove empty rows
    for row in reversed(range(1, hoja_datos_finales.max_row + 1)):
        if all(cell.value is None for cell in hoja_datos_finales[row]):
            hoja_datos_finales.delete_rows(row)       


    # Guardar los cambios en el archivo de Excel
    libro.save(ruta_excel)


def main():
    archivo_excel = seleccionar_archivo_excel()
    nombre_hoja = 'Sheet1'
    if archivo_excel:
        df_tabla = analizar_excel(archivo_excel, 'G:S')
        ruta = exportar_tabla(df_tabla, archivo_excel)
        df_final = optimizar(ruta,nombre_hoja)
        inyectar_df_en_excel(df_final, ruta)
    else:
        print('Nada es seguro, solo el presente')

if __name__ == "__main__":
    main()
